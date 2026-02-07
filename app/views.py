"""
API视图
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.contrib.auth import authenticate, login, logout
from django.views.decorators.csrf import ensure_csrf_cookie
from django.utils.decorators import method_decorator
from django.utils import timezone
from django.http import HttpResponse, StreamingHttpResponse, JsonResponse, FileResponse, Http404
from django.views import View
import json
import hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io
import re
import threading
import os
from datetime import datetime
from pathlib import Path
from .models import Plugin, Task, TaskExecution, Asset, AliyunConfig, AWSConfig, Vulnerability, HexStrikeExecution
from django.db import transaction
from django.db.models import Q
from .services.secops_agent import SecOpsAgent
from .services.hexstrike_client import HexStrikeClient
from django.conf import settings as django_settings
from .utils.hexstrike_export import HexStrikeReportExporter
from .serializers import (
    PluginSerializer, TaskSerializer, TaskExecutionSerializer, AssetSerializer,
    AliyunConfigSerializer, AWSConfigSerializer, VulnerabilitySerializer
)
from .tasks import execute_task_async
from .pagination import CustomPageNumberPagination
from .utils.sse_manager import sse_manager
from .utils.dingtalk import send_dingtalk_message
import logging

logger = logging.getLogger(__name__)


@method_decorator(ensure_csrf_cookie, name='dispatch')
class LoginView(APIView):
    """登录视图"""
    permission_classes = [AllowAny]
    
    def get(self, request):
        """获取CSRF token"""
        return Response({'message': 'CSRF token已设置'})
    
    def post(self, request):
        """用户登录"""
        username = request.data.get('username')
        password = request.data.get('password')
        
        if not username or not password:
            return Response({'error': '用户名和密码不能为空'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 记录登录尝试
        logger.info(f'登录尝试: 用户名={username}')
        
        # 先检查用户是否存在
        try:
            from django.contrib.auth.models import User
            user_obj = User.objects.get(username=username)
            logger.info(f'用户存在: ID={user_obj.id}, is_active={user_obj.is_active}, is_superuser={user_obj.is_superuser}')
            if not user_obj.is_active:
                logger.warning(f'用户 {username} 已被禁用')
                return Response({'error': '用户已被禁用'}, status=status.HTTP_401_UNAUTHORIZED)
        except User.DoesNotExist:
            logger.warning(f'用户不存在: {username}')
            return Response({'error': '用户名或密码错误'}, status=status.HTTP_401_UNAUTHORIZED)
        except Exception as e:
            logger.error(f'检查用户时出错: {str(e)}')
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            logger.info(f'登录成功: 用户名={username}, ID={user.id}')
            return Response({
                'message': '登录成功',
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email
                }
            })
        else:
            logger.warning(f'密码验证失败: 用户名={username}')
            return Response({'error': '用户名或密码错误'}, status=status.HTTP_401_UNAUTHORIZED)


class LogoutView(APIView):
    """登出视图"""
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        logout(request)
        return Response({'message': '登出成功'})


class CurrentUserView(APIView):
    """获取当前用户信息"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user
        return Response({
            'id': user.id,
            'username': user.username,
            'email': user.email
        })


class PluginViewSet(viewsets.ReadOnlyModelViewSet):
    """插件视图集"""
    permission_classes = [IsAuthenticated]
    queryset = Plugin.objects.filter(is_active=True)
    serializer_class = PluginSerializer
    
    @action(detail=True, methods=['get'])
    def tasks(self, request, pk=None):
        """获取插件关联的任务"""
        plugin = self.get_object()
        tasks = Task.objects.filter(plugin=plugin)
        serializer = TaskSerializer(tasks, many=True)
        return Response(serializer.data)


class TaskViewSet(viewsets.ModelViewSet):
    """任务视图集"""
    permission_classes = [IsAuthenticated]
    queryset = Task.objects.all()
    serializer_class = TaskSerializer
    pagination_class = CustomPageNumberPagination
    
    def get_queryset(self):
        # 超级管理员可查看全部任务；普通用户仅查看自己创建的任务
        if getattr(self.request.user, 'is_superuser', False):
            return Task.objects.all()
        return Task.objects.filter(created_by=self.request.user.username)
    
    def create(self, request, *args, **kwargs):
        """创建任务"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        task = serializer.save(created_by=request.user.username)
        
        # 如果是定时任务（cron或interval），自动注册到Celery Beat
        if task.cron_expression and task.is_active and task.trigger_type in ['cron', 'interval']:
            try:
                from app.schedulers import register_task_schedule, restart_celery_beat
                success = register_task_schedule(task)
                if success:
                    trigger_type_name = '间隔任务' if task.trigger_type == 'interval' else '定时任务'
                    logger.info(f"已自动注册{trigger_type_name}到Celery Beat: {task.name} (ID: {task.id})")
                    # 自动重启Celery Beat使新任务生效
                    logger.info(f"正在自动重启Celery Beat以使新任务生效...")
                    restart_success = restart_celery_beat()
                    if restart_success:
                        logger.info(f"✅ Celery Beat已自动重启，新任务已生效: {task.name} (ID: {task.id})")
                    else:
                        logger.warning(f"⚠️  Celery Beat自动重启失败，请手动重启: pkill -f 'celery.*beat' && celery -A app.celery_app beat -l info")
            except Exception as e:
                logger.error(f"自动注册定时任务失败: {task.name} (ID: {task.id}) - {str(e)}", exc_info=True)
        
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        """更新任务"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        
        # 保存旧的属性值，用于判断是否需要重新注册
        old_trigger_type = instance.trigger_type
        old_cron_expression = instance.cron_expression
        old_is_active = instance.is_active
        
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        task = serializer.save()
        
        # 如果定时任务相关属性发生变化，需要重新注册或取消注册
        trigger_changed = (
            task.trigger_type != old_trigger_type or
            task.cron_expression != old_cron_expression or
            task.is_active != old_is_active
        )
        
        if trigger_changed:
            try:
                from app.schedulers import register_task_schedule, unregister_task_schedule, restart_celery_beat
                # 先取消注册旧的（无论新类型是什么，先清理旧的调度）
                unregister_task_schedule(task.id)
                
                # 判断是否需要重启Celery Beat
                need_restart = False
                
                # 如果是从定时改为手动，明确记录
                if old_trigger_type in ['cron', 'interval'] and task.trigger_type == 'manual':
                    logger.info(f"任务已从定时执行改为手动执行，已取消注册定时调度: {task.name} (ID: {task.id})")
                    need_restart = True
                # 如果是从手动改为定时，需要重启
                elif old_trigger_type == 'manual' and task.trigger_type in ['cron', 'interval']:
                    need_restart = True
                # 如果是定时任务（cron或interval）且启用，重新注册
                elif task.cron_expression and task.is_active and task.trigger_type in ['cron', 'interval']:
                    success = register_task_schedule(task)
                    if success:
                        trigger_type_name = '间隔任务' if task.trigger_type == 'interval' else '定时任务'
                        logger.info(f"已自动重新注册{trigger_type_name}到Celery Beat: {task.name} (ID: {task.id})")
                        need_restart = True
                else:
                    logger.info(f"已取消注册定时任务: {task.name} (ID: {task.id}), 新触发类型: {task.trigger_type}")
                    # 如果是从定时改为其他类型，也需要重启
                    if old_trigger_type in ['cron', 'interval']:
                        need_restart = True
                
                # 如果需要重启Celery Beat，自动重启
                if need_restart:
                    logger.info(f"正在自动重启Celery Beat以使任务配置生效...")
                    restart_success = restart_celery_beat()
                    if restart_success:
                        logger.info(f"✅ Celery Beat已自动重启，任务配置已生效: {task.name} (ID: {task.id})")
                    else:
                        logger.warning(f"⚠️  Celery Beat自动重启失败，请手动重启: pkill -f 'celery.*beat' && celery -A app.celery_app beat -l info")
            except Exception as e:
                logger.error(f"自动注册/取消注册定时任务失败: {task.name} (ID: {task.id}) - {str(e)}", exc_info=True)
        
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def execute(self, request, pk=None):
        """手动执行任务"""
        task = self.get_object()
        
        if task.status == 'running':
            return Response(
                {'error': '任务正在执行中，请稍后再试'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 如果没有指定阿里云配置，尝试使用默认配置
        if not task.aliyun_config:
            try:
                default_config = AliyunConfig.objects.filter(
                    user=request.user,
                    is_default=True,
                    is_active=True
                ).first()
                if default_config:
                    task.aliyun_config = default_config
                    task.save(update_fields=['aliyun_config'])
            except Exception as e:
                logger.warning(f'自动关联默认阿里云配置失败: {str(e)}')

        # 创建执行记录，将执行用户ID存储到result字段中（作为元数据）
        execution = TaskExecution.objects.create(
            task=task,
            status='running',
            result={'executor_user_id': request.user.id}  # 存储执行用户ID
        )
        
        # 更新任务状态为运行中（在execute_task_async中也会更新，但这里先更新以便前端立即看到）
        task.status = 'running'
        task.save(update_fields=['status'])
        
        # 发送状态更新通知（向当前执行用户和任务创建者都发送）
        try:
            from app.utils.sse_manager import sse_manager
            from django.contrib.auth.models import User
            
            task_data = {
                'id': task.id,
                'name': task.name,
                'status': 'running',
                'last_run_time': task.last_run_time.isoformat() if task.last_run_time else None,
            }
            
            # 向当前执行用户发送
            sse_manager.send_task_update(request.user.id, task.id, task_data)
            logger.info(f'已向执行用户发送任务状态更新: user_id={request.user.id}, task_id={task.id}')
            
            # 如果任务创建者不同，也向创建者发送
            if task.created_by and task.created_by != request.user.username:
                try:
                    creator = User.objects.filter(username=task.created_by).first()
                    if creator:
                        sse_manager.send_task_update(creator.id, task.id, task_data)
                        logger.info(f'已向任务创建者发送任务状态更新: user_id={creator.id}, task_id={task.id}')
                except Exception as e:
                    logger.warning(f'向任务创建者发送消息失败: {e}')
        except Exception as e:
            logger.warning(f'发送任务状态更新SSE消息失败: {e}')

        # 执行任务（如果celery可用则异步，否则使用线程异步执行）
        try:
            # 尝试异步执行
            from celery import current_app
            if current_app.conf.task_always_eager:
                # Celery配置为同步模式，使用线程异步执行，避免阻塞HTTP请求
                thread = threading.Thread(
                    target=execute_task_async,
                    args=(task.id, execution.id),
                    daemon=True
                )
                thread.start()
                logger.info(f'任务 {task.id} 已在后台线程中开始执行')
            else:
                # 异步执行
                execute_task_async.delay(task.id, execution.id)
                logger.info(f'任务 {task.id} 已提交到Celery异步执行')
        except ImportError:
            # Celery未安装，使用线程异步执行
            thread = threading.Thread(
                target=execute_task_async,
                args=(task.id, execution.id),
                daemon=True
            )
            thread.start()
            logger.info(f'任务 {task.id} 已在后台线程中开始执行（Celery未安装）')
        except Exception as e:
            # 其他异常，也使用线程异步执行
            thread = threading.Thread(
                target=execute_task_async,
                args=(task.id, execution.id),
                daemon=True
            )
            thread.start()
            logger.warning(f'任务 {task.id} 已在后台线程中开始执行（异常：{str(e)}）')
        
        return Response({
            'message': '任务已开始执行',
            'execution_id': execution.id
        })

    @action(detail=True, methods=['get'])
    def executions(self, request, pk=None):
        """获取任务的执行历史"""
        task = self.get_object()
        executions = TaskExecution.objects.filter(task=task).order_by('-started_at')
        serializer = TaskExecutionSerializer(executions, many=True)
        return Response(serializer.data)


class TaskSSEView(APIView):
    """任务状态SSE推送视图"""
    permission_classes = [IsAuthenticated]
    
    def perform_content_negotiation(self, request, force=False):
        """重写内容协商，对于SSE请求直接返回None，跳过协商"""
        # 如果是SSE请求（Accept头包含text/event-stream），跳过内容协商
        accept = request.META.get('HTTP_ACCEPT', '')
        if 'text/event-stream' in accept:
            return (None, None)
        return super().perform_content_negotiation(request, force)
    
    def dispatch(self, request, *args, **kwargs):
        """重写dispatch，直接处理SSE请求，绕过内容协商"""
        # 检查认证
        self.check_permissions(request)
        self.check_throttles(request)
        
        # 直接调用get方法
        response = self.get(request, *args, **kwargs)
        
        # 如果是StreamingHttpResponse，直接返回，不经过finalize_response
        if isinstance(response, StreamingHttpResponse):
            return response
        
        # 否则使用默认的finalize_response
        return self.finalize_response(request, response, *args, **kwargs)
    
    def get(self, request):
        """建立SSE连接，推送任务状态更新"""
        user_id = request.user.id
        logger.info(f'建立SSE连接: user_id={user_id}, username={request.user.username}')
        
        # 将user_id保存到局部变量，避免闭包问题
        current_user_id = user_id
        
        def event_stream():
            """SSE事件流生成器"""
            try:
                # 先发送一个连接成功的消息
                connected_msg = json.dumps({'type': 'connected', 'message': 'SSE连接已建立'}, ensure_ascii=False)
                yield f"data: {connected_msg}\n\n"
                logger.info(f'SSE连接成功消息已发送: user_id={current_user_id}')
                
                # 持续监听消息
                # 使用较短的超时时间，以便更快地发送心跳和检查新消息
                while True:
                    try:
                        # 从消息队列获取消息（使用较短的超时时间，以便更快响应）
                        message = sse_manager.get_message(current_user_id, timeout=5.0)
                        if message:
                            # 检查是否是心跳消息，心跳消息使用debug级别
                            try:
                                msg_data = json.loads(message)
                                if msg_data.get('type') == 'heartbeat':
                                    logger.debug(f'SSE发送心跳给用户 {current_user_id}')
                                else:
                                    logger.info(f'SSE发送消息给用户 {current_user_id}: {message[:200]}')
                            except:
                                logger.info(f'SSE发送消息给用户 {current_user_id}: {message[:200]}')
                            yield f"data: {message}\n\n"
                    except Exception as msg_error:
                        logger.error(f'获取SSE消息时出错: {msg_error}', exc_info=True)
                        # 继续循环，不要中断连接
                        continue
            except GeneratorExit:
                # 客户端断开连接
                logger.info(f'SSE连接已断开: user_id={current_user_id}')
                raise
            except Exception as e:
                logger.error(f'SSE流异常: {e}', exc_info=True)
                import traceback
                logger.error(f'SSE流异常Traceback: {traceback.format_exc()}')
                try:
                    error_msg = json.dumps({'type': 'error', 'message': str(e)}, ensure_ascii=False)
                    yield f"data: {error_msg}\n\n"
                except:
                    pass  # 如果连错误消息都发不出去，就放弃
            finally:
                # 清理队列
                try:
                    sse_manager.clear_queue(current_user_id)
                    logger.info(f'已清理SSE队列: user_id={current_user_id}')
                except Exception as cleanup_error:
                    logger.error(f'清理SSE队列失败: {cleanup_error}')
        
        response = StreamingHttpResponse(
            event_stream(),
            content_type='text/event-stream'
        )
        response['Cache-Control'] = 'no-cache'
        response['X-Accel-Buffering'] = 'no'
        return response


class TaskExecutionViewSet(viewsets.ReadOnlyModelViewSet):
    """任务执行记录视图集"""
    permission_classes = [IsAuthenticated]
    queryset = TaskExecution.objects.all()
    serializer_class = TaskExecutionSerializer
    pagination_class = CustomPageNumberPagination


class AssetViewSet(viewsets.ReadOnlyModelViewSet):
    """资产视图集"""
    permission_classes = [IsAuthenticated]
    queryset = Asset.objects.all()
    serializer_class = AssetSerializer
    pagination_class = CustomPageNumberPagination
    
    def get_queryset(self):
        queryset = Asset.objects.all()
        
        # 筛选条件
        source = self.request.query_params.get('source', None)
        asset_type = self.request.query_params.get('asset_type', None)
        name = self.request.query_params.get('name', None)
        uuid = self.request.query_params.get('uuid', None)
        host_uuid = self.request.query_params.get('host_uuid', None)
        intranet_ip = self.request.query_params.get('intranet_ip', None)
        internet_ip = self.request.query_params.get('internet_ip', None)
        
        if source:
            queryset = queryset.filter(source=source)
        
        if asset_type:
            # 支持多个类型筛选
            if isinstance(asset_type, str):
                asset_types = [asset_type]
            else:
                asset_types = asset_type if isinstance(asset_type, list) else [asset_type]
            queryset = queryset.filter(asset_type__in=asset_types)
        
        if name:
            queryset = queryset.filter(name__icontains=name)
        
        if uuid:
            queryset = queryset.filter(uuid__icontains=uuid)
        
        if host_uuid:
            queryset = queryset.filter(host_uuid__icontains=host_uuid)
        
        if intranet_ip or internet_ip:
            # IP地址可能在data字段中
            ip_query = Q()
            if intranet_ip:
                ip_query |= Q(data__icontains=intranet_ip)
            if internet_ip:
                ip_query |= Q(data__icontains=internet_ip)
            if ip_query:
                queryset = queryset.filter(ip_query)
        
        return queryset.order_by('-collected_at')
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """导出资产数据到Excel"""
        queryset = self.get_queryset()
        
        # 创建Excel工作簿
        wb = Workbook()
        
        # 按资产类型分组导出
        asset_types = queryset.values_list('asset_type', flat=True).distinct()
        type_name_map = {
            'server': '服务器',
            'account': '账户',
            'port': '端口',
            'process': '进程',
            'middleware': '中间件',
            'software': '软件',
            'ai_component': 'AI组件',
            'cron_task': '计划任务',
            'startup_item': '启动项',
            'kernel_module': '内核模块',
            'web_site': 'Web站点',
            'idc_probe': 'IDC探针发现',
            'database': '数据库',
            'web_service': 'Web服务'
        }
        
        for asset_type in asset_types:
            type_assets = queryset.filter(asset_type=asset_type)
            type_name = type_name_map.get(asset_type, asset_type)
            ws = wb.create_sheet(title=type_name[:31])  # Excel工作表名称最多31个字符
            
            # 基础字段
            base_fields = ['资产类型', '名称', 'UUID', '主机UUID', '数据来源', '采集时间', '更新时间']
            headers = base_fields.copy()
            
            # 动态字段映射（中文-英文）
            field_mapping = {}
            all_keys = set()
            for asset in type_assets:
                if isinstance(asset.data, dict):
                    all_keys.update(asset.data.keys())
            
            # 为每个key添加中文映射（简化版，实际可以根据需要完善）
            for key in sorted(all_keys):
                if key not in ['asset_type', 'uuid', 'host_uuid', 'name', 'source']:
                    chn_name = key.replace('_', ' ').title()
                    field_mapping[key] = chn_name
                    headers.append(chn_name)
            
            # 写入表头
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 写入数据
            for row_idx, asset in enumerate(type_assets, 2):
                row_data = [
                    type_name_map.get(asset.asset_type, asset.asset_type),
                    asset.name or '',
                    asset.uuid or '',
                    asset.host_uuid or '',
                    asset.source or '',
                    asset.collected_at.strftime('%Y-%m-%d %H:%M:%S') if asset.collected_at else '',
                    asset.updated_at.strftime('%Y-%m-%d %H:%M:%S') if asset.updated_at else ''
                ]
                
                # 获取data字段
                asset_data = asset.data if isinstance(asset.data, dict) else {}
                
                # 填充data字段的值
                for header in headers[len(base_fields):]:
                    # 找到对应的英文key
                    english_key = None
                    for eng, chn in field_mapping.items():
                        if chn == header:
                            english_key = eng
                            break
                    if not english_key:
                        english_key = header
                    
                    value = asset_data.get(english_key, '')
                    # 处理不同类型的值
                    if isinstance(value, bool):
                        value = '是' if value else '否'
                    elif isinstance(value, (int, float)) and english_key in ['CreateTimestamp', 'ProcessStarted', 'InstallTimeDt', 'LastLoginTimestamp']:
                        # 时间戳转换为日期时间
                        try:
                            if value > 1000000000000:  # 毫秒时间戳
                                value = datetime.fromtimestamp(value / 1000).strftime('%Y-%m-%d %H:%M:%S')
                            else:  # 秒时间戳
                                value = datetime.fromtimestamp(value).strftime('%Y-%m-%d %H:%M:%S')
                        except:
                            pass
                    elif isinstance(value, (list, tuple)):
                        value = ', '.join(str(v) for v in value)
                    elif isinstance(value, dict):
                        value = json.dumps(value, ensure_ascii=False)
                    row_data.append(str(value) if value is not None else '')
                
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col, value=value)
            
            # 调整列宽
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].width = 15
        
        # 删除默认的Sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 创建响应
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'资产数据导出_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class AWSConfigViewSet(viewsets.ModelViewSet):
    """AWS配置视图集"""
    permission_classes = [IsAuthenticated]
    serializer_class = AWSConfigSerializer
    pagination_class = CustomPageNumberPagination
    
    def get_queryset(self):
        return AWSConfig.objects.filter(user=self.request.user)
    
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
    
    def update(self, request, *args, **kwargs):
        """更新配置"""
        return super().update(request, *args, **kwargs)


class VulnerabilityViewSet(viewsets.ReadOnlyModelViewSet):
    """漏洞视图集"""
    permission_classes = [IsAuthenticated]
    queryset = Vulnerability.objects.all()
    serializer_class = VulnerabilitySerializer
    pagination_class = CustomPageNumberPagination
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    
    def get_queryset(self):
        queryset = Vulnerability.objects.all()
        
        source = self.request.query_params.get('source', None)
        cve_id = self.request.query_params.get('cve_id', None)
        title = self.request.query_params.get('title', None)
        published_date_from = self.request.query_params.get('published_date_from', None)
        published_date_to = self.request.query_params.get('published_date_to', None)
        
        if source:
            queryset = queryset.filter(source=source)
        
        if cve_id:
            queryset = queryset.filter(cve_id__icontains=cve_id)
        
        if title:
            queryset = queryset.filter(title__icontains=title)
        
        if published_date_from:
            queryset = queryset.filter(published_date__gte=published_date_from)
        
        if published_date_to:
            queryset = queryset.filter(published_date__lte=published_date_to)
        
        return queryset.order_by('-published_date', '-collected_at')
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """导出漏洞数据到Excel"""
        queryset = self.get_queryset()
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = '漏洞数据'
        
        # 表头
        headers = ['CVE编号', '标题', '发布日期', '数据来源', '消息ID', '采集时间', '更新时间', '邮件链接']
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 数据
        for row_idx, vuln in enumerate(queryset, 2):
            row_data = [
                vuln.cve_id,
                vuln.title,
                vuln.published_date.strftime('%Y-%m-%d') if vuln.published_date else '',
                vuln.source,
                vuln.message_id,
                vuln.collected_at.strftime('%Y-%m-%d %H:%M:%S') if vuln.collected_at else '',
                vuln.updated_at.strftime('%Y-%m-%d %H:%M:%S') if vuln.updated_at else '',
                vuln.url
            ]
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=value)
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 20
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 创建响应
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'漏洞数据导出_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @action(detail=False, methods=['post'])
    def import_cnvd_xml(self, request):
        """手动导入 CNVD 格式的 XML 文件，解析并存储为漏洞数据（source=cnvd）。"""
        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response(
                {'error': '请上传文件', 'detail': '请选择 XML 文件（form-data 字段名: file）'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not uploaded.name.lower().endswith('.xml'):
            return Response(
                {'error': '仅支持 XML 文件', 'detail': '请上传 .xml 后缀的文件'},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            raw = uploaded.read()
            content = raw.decode('utf-8')
        except UnicodeDecodeError:
            try:
                content = raw.decode('utf-8-sig')
            except Exception as e:
                return Response(
                    {'error': '文件编码错误', 'detail': str(e)},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            return Response(
                {'error': '读取文件失败', 'detail': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            from .utils.cnvd_xml_parser import parse_cnvd_xml
            items = parse_cnvd_xml(content)
        except Exception as e:
            return Response(
                {'error': 'XML 解析失败', 'detail': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not items:
            return Response(
                {'error': '未解析到漏洞数据', 'detail': '请确认 XML 根节点为 vulnerabilitys，子节点为 vulnerability'},
                status=status.HTTP_400_BAD_REQUEST
            )
        created = 0
        updated = 0
        cnvd_base_url = 'https://www.cnvd.org.cn/flaw/show/'
        with transaction.atomic():
            for item in items:
                cnvd_id = item.get('cnvd_id') or ''
                title = (item.get('title') or cnvd_id)[:500]
                description = item.get('description') or ''
                published_date = item.get('published_date')
                content_obj = item.get('content') or {}
                url = f'{cnvd_base_url}{cnvd_id}' if cnvd_id else ''
                if not url:
                    continue
                existing = Vulnerability.objects.filter(source='cnvd', url=url).first()
                if not existing:
                    existing = Vulnerability.objects.filter(source='cnvd', cve_id=cnvd_id).first()
                if existing:
                    existing.title = title
                    existing.description = description
                    existing.content = {**(existing.content or {}), **content_obj}
                    if published_date:
                        existing.published_date = published_date
                    existing.raw_content = (description or title)[:50000]
                    existing.save()
                    updated += 1
                else:
                    Vulnerability.objects.create(
                        cve_id=cnvd_id,
                        title=title,
                        description=description,
                        url=url,
                        message_id='',
                        published_date=published_date,
                        raw_content=(description or title)[:50000],
                        content=content_obj,
                        source='cnvd'
                    )
                    created += 1
        return Response({
            'message': '导入完成',
            'total': len(items),
            'created': created,
            'updated': updated,
        })
    
    @action(detail=True, methods=['post'])
    def send_to_dingtalk(self, request, pk=None):
        """发送漏洞信息到钉钉"""
        vulnerability = self.get_object()
        
        # 获取用户的钉钉配置
        dingtalk_config = AliyunConfig.objects.filter(
            user=request.user,
            is_active=True
        ).filter(
            Q(config_type='dingtalk') | Q(config_type='both')
        ).filter(
            dingtalk_enabled=True
        ).exclude(
            dingtalk_webhook=''
        ).first()
        
        if not dingtalk_config or not dingtalk_config.dingtalk_webhook:
            return Response(
                {'error': '未找到可用的钉钉配置，请先在系统配置中配置钉钉机器人'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 构建消息内容
        content = vulnerability.content if isinstance(vulnerability.content, dict) else {}
        
        title = f"漏洞提醒: {vulnerability.cve_id}"
        text = f"## {title}\n\n"
        text += f"**CVE编号**: {vulnerability.cve_id}\n\n"
        text += f"**标题**: {vulnerability.title}\n\n"
        
        if content.get('basic_description'):
            text += f"**基本描述**: {content['basic_description']}\n\n"
        elif vulnerability.description:
            text += f"**描述**: {vulnerability.description[:200]}...\n\n"
        
        if content.get('severity'):
            text += f"**危害等级**: {content['severity']}\n\n"
        
        if content.get('affected_component'):
            text += f"**影响组件**: {content['affected_component']}\n\n"
        
        if content.get('affected_versions'):
            text += f"**受影响版本**: {content['affected_versions']}\n\n"
        
        if content.get('impact'):
            text += f"**影响**: {content['impact'][:300]}...\n\n"
        
        if content.get('solution'):
            text += f"**解决方案**: {content['solution'][:300]}...\n\n"
        
        text += f"**详情链接**: {vulnerability.url}\n"
        
        # 发送消息
        result = send_dingtalk_message(
            webhook_url=dingtalk_config.dingtalk_webhook,
            secret=dingtalk_config.dingtalk_secret if dingtalk_config.dingtalk_secret else None,
            title=title,
            text=text
        )
        
        if result['success']:
            return Response({'message': '漏洞信息已发送到钉钉'})
        else:
            return Response(
                {'error': f"发送失败: {result['message']}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def send_to_feishu(self, request, pk=None):
        """发送漏洞信息到飞书"""
        from .utils.feishu import send_vulnerability_to_feishu
        vulnerability = self.get_object()
        
        # 获取用户的飞书配置
        feishu_config = AliyunConfig.objects.filter(
            user=request.user,
            is_active=True
        ).filter(
            Q(config_type='feishu') | Q(config_type='both')
        ).filter(
            feishu_enabled=True
        ).exclude(
            feishu_webhook=''
        ).first()
        
        if not feishu_config or not feishu_config.feishu_webhook:
            return Response(
                {'error': '未找到可用的飞书配置，请先在系统配置中配置飞书机器人'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 使用工具函数发送消息
        result = send_vulnerability_to_feishu(feishu_config, vulnerability)
        
        if result['success']:
            return Response({'message': '漏洞信息已发送到飞书'})
        else:
            return Response(
                {'error': f"发送失败: {result['message']}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class AliyunConfigViewSet(viewsets.ModelViewSet):
    """系统配置视图集"""
    permission_classes = [IsAuthenticated]
    serializer_class = AliyunConfigSerializer
    pagination_class = CustomPageNumberPagination
    
    def get_queryset(self):
        return AliyunConfig.objects.filter(user=self.request.user)
    
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
    
    def update(self, request, *args, **kwargs):
        """更新配置"""
        return super().update(request, *args, **kwargs)
    
    @action(detail=True, methods=['post'])
    def test_qianwen(self, request, pk=None):
        """测试通义千问API连接"""
        from .utils.qianwen import test_qianwen_connection
        config = self.get_object()
        if not config.qianwen_api_key:
            return Response({'error': '通义千问API Key未配置'}, status=status.HTTP_400_BAD_REQUEST)
        result = test_qianwen_connection(
            api_key=config.qianwen_api_key,
            api_base=config.qianwen_api_base or 'https://dashscope.aliyuncs.com/compatible-mode/v1',
            model=config.qianwen_model or 'qwen-plus'
        )
        if result['success']:
            return Response({'message': result['message'], 'data': result.get('data', {})})
        else:
            return Response({'error': result['message']}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def test_ai_parsing(self, request, pk=None):
        """测试AI解析功能（使用示例漏洞内容）"""
        from .utils.qianwen import parse_vulnerability_with_ai
        config = self.get_object()
        if not config.qianwen_api_key or not config.qianwen_enabled:
            return Response({'error': '通义千问配置未启用或API Key未配置'}, status=status.HTTP_400_BAD_REQUEST)
        sample_content = """Subject: CVE-2025-12345: Apache Example: Security vulnerability

Severity: High

Affected versions:
- Apache Example 1.0.0 through 1.2.5

Description:
A security vulnerability has been discovered in Apache Example that allows
remote attackers to execute arbitrary code.

Impact:
An attacker may be able to execute arbitrary code on the server.

Solution:
Users are recommended to upgrade to version 1.3.0 or higher, which fixes the issue.

References:
https://example.com/security/CVE-2025-12345
"""
        try:
            result = parse_vulnerability_with_ai(
                raw_content=sample_content,
                cve_id='CVE-2025-12345',
                api_key=config.qianwen_api_key,
                api_base=config.qianwen_api_base or 'https://dashscope.aliyuncs.com/compatible-mode/v1',
                model=config.qianwen_model or 'qwen-plus'
            )
            if result and isinstance(result, dict) and len(result) > 0:
                return Response({'message': 'AI解析测试成功', 'data': result})
            else:
                logger.warning(f"AI解析返回空结果，result类型: {type(result)}, 内容: {result}")
                return Response(
                    {'error': 'AI解析返回空结果', 'detail': '可能的原因：1. API Key无效 2. 模型不支持JSON格式输出 3. API调用失败。请查看后端日志获取详细信息。'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            error_msg = str(e)
            logger.error(f"AI解析测试失败: {error_msg}", exc_info=True)
            return Response(
                {'error': f'AI解析测试失败: {error_msg}', 'detail': '请检查：1. API Key是否正确 2. API地址是否正确 3. 网络连接是否正常 4. 模型名称是否正确'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def test_dingtalk(self, request, pk=None):
        """测试钉钉配置"""
        config = self.get_object()
        if not config.dingtalk_webhook or not config.dingtalk_enabled:
            return Response({'error': '钉钉配置未启用或Webhook未配置'}, status=status.HTTP_400_BAD_REQUEST)
        
        result = send_dingtalk_message(
            webhook_url=config.dingtalk_webhook,
            secret=config.dingtalk_secret if config.dingtalk_secret else None,
            title='测试消息',
            text='这是一条测试消息，用于验证钉钉机器人配置是否正确。'
        )
        
        if result['success']:
            return Response({'message': '测试消息发送成功'})
        else:
            return Response({'error': f"发送失败: {result['message']}"}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def test_dingtalk_config(self, request):
        """测试钉钉配置（支持未保存的配置）"""
        webhook_url = request.data.get('dingtalk_webhook', '').strip()
        secret = request.data.get('dingtalk_secret', '').strip() or None
        
        if not webhook_url:
            return Response({'error': '钉钉Webhook地址不能为空'}, status=status.HTTP_400_BAD_REQUEST)
        
        result = send_dingtalk_message(
            webhook_url=webhook_url,
            secret=secret,
            title='测试消息',
            text='这是一条测试消息，用于验证钉钉机器人配置是否正确。'
        )
        
        if result['success']:
            return Response({'message': '测试消息发送成功'})
        else:
            return Response({'error': f"发送失败: {result['message']}"}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def test_feishu(self, request, pk=None):
        """测试飞书配置"""
        from .utils.feishu import send_feishu_message
        config = self.get_object()
        if not config.feishu_webhook or not config.feishu_enabled:
            return Response({'error': '飞书配置未启用或Webhook未配置'}, status=status.HTTP_400_BAD_REQUEST)
        
        result = send_feishu_message(
            webhook_url=config.feishu_webhook,
            secret=config.feishu_secret if config.feishu_secret else None,
            title='测试消息',
            text='这是一条测试消息，用于验证飞书机器人配置是否正确。'
        )
        
        if result['success']:
            return Response({'message': '测试消息发送成功'})
        else:
            return Response({'error': f"发送失败: {result['message']}"}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def test_feishu_config(self, request):
        """测试飞书配置（支持未保存的配置）"""
        from .utils.feishu import send_feishu_message
        webhook_url = request.data.get('feishu_webhook', '').strip()
        secret = request.data.get('feishu_secret', '').strip() or None
        
        if not webhook_url:
            return Response({'error': '飞书Webhook地址不能为空'}, status=status.HTTP_400_BAD_REQUEST)
        
        result = send_feishu_message(
            webhook_url=webhook_url,
            secret=secret,
            title='测试消息',
            text='这是一条测试消息，用于验证飞书机器人配置是否正确。'
        )
        
        if result['success']:
            return Response({'message': '测试消息发送成功'})
        else:
            return Response({'error': f"发送失败: {result['message']}"}, status=status.HTTP_400_BAD_REQUEST)


class SecOpsAgentViewSet(viewsets.ViewSet):
    """SecOps智能体视图集"""
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['post'])
    def chat(self, request):
        """
        与智能体对话，流式返回响应
        
        Request Body:
            {
                "message": "用户消息",
                "conversation_history": [{"role": "user", "content": "..."}, ...]  # 可选
            }
        """
        try:
            user_message = request.data.get('message', '')
            logger.info(
                "SecOps chat 请求: message_len=%d, preview=%s",
                len(user_message or ''),
                (user_message or '')[:100],
            )
            if not user_message:
                return Response({'error': '消息不能为空'}, status=status.HTTP_400_BAD_REQUEST)

            # 视图层兜底：若消息为「安全评估 + IP/域名」，直接调用 HexStrike，不依赖 agent.chat() 路径
            ip_match = re.search(r'(?:\d{1,3}\.){3}\d{1,3}', user_message)
            domain_match = re.search(
                r'(?:[a-zA-Z0-9](?:[a-zA-Z0-9-]*[a-zA-Z0-9])?\.)+[a-zA-Z]{2,}',
                user_message,
            )
            hexstrike_target = (ip_match.group(0) if ip_match else None) or (domain_match.group(0) if domain_match else None)
            security_keywords = [
                '安全评估', '渗透测试', '漏洞扫描', '全面评估', '全面的安全评估', '全面安全评估',
                '安全扫描', '扫描一下', '做一次评估', '做一次扫描', '评估', '扫描',
            ]
            has_security_keyword = any(kw in user_message for kw in security_keywords)
            has_asset_keyword = any(kw in user_message for kw in ['资产', '服务器', '目标', '对', '云服务器'])
            has_security_intent = has_security_keyword or (hexstrike_target and has_asset_keyword)
            hexstrike_enabled = getattr(django_settings, 'HEXSTRIKE_ENABLED', True)
            
            matched_security_keywords = [kw for kw in security_keywords if kw in user_message]
            matched_asset_keywords = [kw for kw in ['资产', '服务器', '目标', '对', '云服务器'] if kw in user_message]
            
            logger.info(
                "SecOps 视图层意图检测: hexstrike_target=%s, has_security_keyword=%s (%s), has_asset_keyword=%s (%s), "
                "has_security_intent=%s, hexstrike_enabled=%s, user_message_preview=%s",
                hexstrike_target, has_security_keyword, matched_security_keywords, 
                has_asset_keyword, matched_asset_keywords, has_security_intent, 
                hexstrike_enabled, (user_message or '')[:100]
            )
            
            if has_security_intent and hexstrike_target and hexstrike_enabled:
                logger.info("✓ SecOps 视图层直接调用 HexStrike: target=%s", hexstrike_target)

                def generate_hexstrike_response():
                    base_url = getattr(django_settings, 'HEXSTRIKE_SERVER_URL', 'http://localhost:8888')
                    timeout = getattr(django_settings, 'HEXSTRIKE_TIMEOUT', 600)  # 增加到 10 分钟
                    client = HexStrikeClient(base_url=base_url, timeout=timeout)

                    # 1) AI 目标分析（返回目标画像和策略建议）
                    result = client.analyze_target(hexstrike_target, analysis_type='comprehensive')
                    logger.info("HexStrike analyze_target 结果: success=%s", result.get('success'))

                    response_parts = []
                    response_parts.append(f'### ✅ 已对目标 {hexstrike_target} 完成安全分析\n\n')

                    if result.get('success') and result.get('data') is not None:
                        data = result['data']

                        # 格式化并显示目标画像
                        if isinstance(data, dict) and 'target_profile' in data:
                            target_profile = data['target_profile']
                            response_parts.append("## 📊 目标画像\n\n")
                            response_parts.append(json.dumps(target_profile, ensure_ascii=False, indent=2))
                            response_parts.append("\n\n")

                    # 2) 执行 nmap 端口扫描
                    logger.info("开始执行 Nmap 端口扫描: target=%s", hexstrike_target)
                    nmap_res = client.run_command("nmap_scan", {"target": hexstrike_target})

                    if nmap_res.get('success') and nmap_res.get('data') is not None:
                        nmap_data = nmap_res['data']
                        stdout = nmap_data.get('stdout', '')
                        stderr = nmap_data.get('stderr', '')

                        # 尝试格式化 Nmap 结果
                        if stdout or stderr:
                            try:
                                from app.services.nmap_result_parser import format_nmap_result
                                formatted_nmap = format_nmap_result(stdout, stderr)
                                response_parts.append("## 🔍 Nmap 端口扫描结果\n\n")
                                response_parts.append(formatted_nmap)
                                response_parts.append("\n\n")
                            except Exception as e:
                                logger.warning(f"Nmap 结果格式化失败: {e}")
                                response_parts.append("## 🔍 Nmap 端口扫描结果\n\n")
                                response_parts.append(f"```\n{stdout[:1000]}\n```\n\n")
                        logger.info("Nmap 扫描成功")
                    elif not nmap_res.get('success'):
                        error_msg = nmap_res.get('message', '未执行或失败')
                        if 'timed out' in error_msg.lower():
                            response_parts.append("## ⏱️ Nmap 端口扫描结果\n\n")
                            response_parts.append("⚠️ 扫描超时\n\n")
                        else:
                            response_parts.append(f"**Nmap**：{error_msg}\n\n")
                        logger.warning("Nmap 扫描失败: %s", nmap_res.get('message'))

                    # 3) 执行 nuclei 漏洞扫描
                    logger.info("开始执行 Nuclei 漏洞扫描: target=%s", hexstrike_target)
                    nuclei_res = client.run_command("nuclei_scan", {"target": hexstrike_target})

                    if nuclei_res.get('success') and nuclei_res.get('data') is not None:
                        nuclei_data = nuclei_res['data']
                        stdout = nuclei_data.get('stdout', '')
                        stderr = nuclei_data.get('stderr', '')

                        # 尝试格式化 Nuclei 结果
                        if stdout or stderr:
                            try:
                                from app.services.nuclei_result_parser import format_nuclei_result
                                formatted_nuclei = format_nuclei_result(stdout, stderr)
                                response_parts.append("## 🔍 Nuclei 漏洞扫描结果\n\n")
                                response_parts.append(formatted_nuclei)
                                response_parts.append("\n\n")
                            except Exception as e:
                                logger.warning(f"Nuclei 结果格式化失败: {e}")
                                response_parts.append("## 🔍 Nuclei 漏洞扫描结果\n\n")
                                response_parts.append(f"```\n{stdout[:1000]}\n```\n\n")
                        logger.info("Nuclei 扫描成功")
                    elif not nuclei_res.get('success'):
                        error_msg = nuclei_res.get('message', '未执行或失败')
                        if 'timed out' in error_msg.lower() or 'timeout' in error_msg.lower():
                            response_parts.append("## ⏱️ Nuclei 漏洞扫描结果\n\n")
                            response_parts.append("⚠️ 扫描超时（超过10分钟），建议分端口扫描或减少扫描范围\n\n")
                        else:
                            response_parts.append(f"**Nuclei**：{error_msg}\n\n")
                        logger.warning("Nuclei 扫描失败: %s", nuclei_res.get('message'))

                    response_parts.append(f"---\n✅ 评估完成。查看 HexStrike 执行过程：`docker logs hexstrike-ai 2>&1 | grep -E \"EXECUTING|FINAL RESULTS|{hexstrike_target}\"`")

                    # 生成 HTML 报告
                    report_filename = None
                    try:
                        from app.services.hexstrike_html_reporter import HexStrikeHTMLReporter
                        reporter = HexStrikeHTMLReporter()

                        # 收集结果数据
                        nmap_data = nmap_res.get('data') if nmap_res else None
                        nuclei_data = nuclei_res.get('data') if nuclei_res else None
                        target_profile = result.get('data', {}).get('target_profile') if result.get('data') else None

                        # 生成报告
                        report_filename = reporter.generate_report(
                            target=hexstrike_target,
                            nmap_results=nmap_data,
                            nuclei_results=nuclei_data,
                            target_profile=target_profile
                        )

                        logger.info(f"HexStrike HTML 报告已生成: {report_filename}")

                        # 在响应末尾添加报告下载链接
                        response_parts.append(f"\n\n📄 **完整报告下载**：[点击下载 HTML 报告](/api/reports/hexstrike/{report_filename})\n")

                    except Exception as e:
                        logger.warning(f"生成 HTML 报告失败: {e}", exc_info=True)

                    # 流式输出响应内容
                    full_response = ''.join(response_parts)
                    # 分块输出，每1000字符一个块，提供更好的流式体验
                    chunk_size = 1000
                    for i in range(0, len(full_response), chunk_size):
                        chunk = full_response[i:i + chunk_size]
                        yield f"data: {json.dumps({'content': chunk}, ensure_ascii=False)}\n\n"
                    
                    yield f"data: {json.dumps({'done': True}, ensure_ascii=False)}\n\n"
                    logger.info("HexStrike 安全评估响应已生成完成")

                response = StreamingHttpResponse(
                    generate_hexstrike_response(),
                    content_type='text/event-stream',
                )
                response['Cache-Control'] = 'no-cache'
                response['X-Accel-Buffering'] = 'no'
                return response
            elif has_security_intent and hexstrike_target and not hexstrike_enabled:
                logger.warning("检测到安全评估意图但 HexStrike 未启用，继续执行 agent.chat()")
            elif has_security_intent and not hexstrike_target:
                logger.warning("检测到安全评估意图但未提取到目标，继续执行 agent.chat()")
            else:
                logger.debug("未检测到安全评估意图，继续执行 agent.chat()")

            conversation_history = request.data.get('conversation_history', [])

            # 获取用户的通义千问配置
            qianwen_config = AliyunConfig.objects.filter(
                user=request.user,
                is_active=True
            ).filter(
                Q(config_type='qianwen') | Q(config_type='both')
            ).filter(
                qianwen_enabled=True
            ).exclude(
                qianwen_api_key=''
            ).first()

            if not qianwen_config or not qianwen_config.qianwen_api_key:
                return Response(
                    {'error': '未找到可用的通义千问配置，请先在系统配置中配置通义千问API'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 创建智能体实例
            try:
                agent = SecOpsAgent(
                    api_key=qianwen_config.qianwen_api_key,
                    api_base=qianwen_config.qianwen_api_base or 'https://dashscope.aliyuncs.com/compatible-mode/v1',
                    model=qianwen_config.qianwen_model or 'qwen-plus'
                )
            except Exception as e:
                logger.error(f"创建智能体失败: {e}", exc_info=True)
                return Response(
                    {'error': f'创建智能体失败: {str(e)}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

            # 流式响应生成器
            def generate_response():
                try:
                    for chunk in agent.chat(user_message, conversation_history, request.user):
                        # 使用SSE格式
                        yield f"data: {json.dumps({'content': chunk}, ensure_ascii=False)}\n\n"
                    # 发送结束标记
                    yield f"data: {json.dumps({'done': True}, ensure_ascii=False)}\n\n"
                except Exception as e:
                    logger.error(f"智能体对话失败: {e}", exc_info=True)
                    error_msg = json.dumps({'error': str(e)}, ensure_ascii=False)
                    yield f"data: {error_msg}\n\n"

            response = StreamingHttpResponse(
                generate_response(),
                content_type='text/event-stream'
            )
            response['Cache-Control'] = 'no-cache'
            response['X-Accel-Buffering'] = 'no'
            return response
        except Exception as e:
            logger.error(f"智能体接口异常: {e}", exc_info=True)
            return Response(
                {'error': f'服务异常: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def hexstrike_status(self, request):
        """
        查询 HexStrike AI 集成状态（是否启用、服务是否可达）。
        用于前端展示「资产安全评估」能力是否可用。
        """
        enabled = getattr(django_settings, 'HEXSTRIKE_ENABLED', True)
        if not enabled:
            return Response({
                'enabled': False,
                'connected': False,
                'message': 'HexStrike 集成未启用（HEXSTRIKE_ENABLED=False）',
                'server_url': getattr(django_settings, 'HEXSTRIKE_SERVER_URL', ''),
            })
        base_url = getattr(django_settings, 'HEXSTRIKE_SERVER_URL', 'http://localhost:8888')
        try:
            client = HexStrikeClient(base_url=base_url, timeout=10)
            result = client.health()
            if result.get('success'):
                return Response({
                    'enabled': True,
                    'connected': True,
                    'message': 'HexStrike 服务正常',
                    'server_url': base_url,
                })
            return Response({
                'enabled': True,
                'connected': False,
                'message': result.get('message', 'HexStrike 服务无响应'),
                'server_url': base_url,
            })
        except Exception as e:
            return Response({
                'enabled': True,
                'connected': False,
                'message': str(e),
                'server_url': base_url,
            })

    @action(detail=False, methods=['get'])
    def hexstrike_reports(self, request):
        """
        获取 HexStrike HTML 报告列表

        查询参数：
        - target: 过滤目标（可选）
        - limit: 返回数量限制（默认 50）
        """
        try:
            from pathlib import Path
            from django.conf import settings

            # 获取查询参数
            target_filter = request.query_params.get('target', '').strip()
            limit = int(request.query_params.get('limit', '50'))

            # 构建 reports 目录路径
            base_dir = Path(settings.BASE_DIR)
            reports_dir = base_dir / 'reports'

            if not reports_dir.exists():
                return Response({
                    'reports': [],
                    'total': 0,
                    'message': '报告目录不存在'
                })

            # 获取所有 HTML 报告文件
            reports = []
            for file_path in sorted(reports_dir.glob('hexstrike_report_*.html'), reverse=True):
                try:
                    # 从文件名提取信息
                    filename = file_path.name
                    stat = file_path.stat()

                    # 解析文件名: hexstrike_report_{target}_{timestamp}.html
                    parts = filename.replace('hexstrike_report_', '').replace('.html', '').rsplit('_', 1)
                    if len(parts) == 2:
                        target_part = parts[0].replace('_', '.')
                        timestamp_part = parts[1]

                        # 应用目标过滤
                        if target_filter and target_filter.lower() not in target_part.lower():
                            continue

                        reports.append({
                            'filename': filename,
                            'target': target_part,
                            'created_at': timestamp_part,
                            'size': stat.st_size,
                            'download_url': f'/api/reports/hexstrike/{filename}',
                            'created_time': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                        })

                        # 限制返回数量
                        if len(reports) >= limit:
                            break

                except Exception as e:
                    logger.warning(f"解析报告文件失败 {filename}: {e}")
                    continue

            return Response({
                'reports': reports,
                'total': len(reports),
                'message': f'找到 {len(reports)} 份报告'
            })

        except Exception as e:
            logger.error(f"获取报告列表失败: {e}", exc_info=True)
            return Response({
                'error': f'获取报告列表失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def hexstrike_export(self, request):
        """
        导出 HexStrike 执行报告
        支持格式：excel, pdf, html
        
        查询参数：
        - format: 导出格式 (excel/pdf/html)，默认 excel
        - target: 评估目标（可选，用于筛选）
        - execution_ids: 执行记录ID列表，逗号分隔（可选）
        - start_date: 开始日期（可选）
        - end_date: 结束日期（可选）
        """
        try:
            format_type = request.query_params.get('format', 'excel').lower()
            target = request.query_params.get('target', None)
            execution_ids_param = request.query_params.get('execution_ids', None)
            start_date = request.query_params.get('start_date', None)
            end_date = request.query_params.get('end_date', None)
            
            # 查询执行记录
            queryset = HexStrikeExecution.objects.all()
            
            if target:
                queryset = queryset.filter(target__icontains=target)
            
            if execution_ids_param:
                execution_ids = [int(id.strip()) for id in execution_ids_param.split(',') if id.strip().isdigit()]
                if execution_ids:
                    queryset = queryset.filter(id__in=execution_ids)
            
            if start_date:
                try:
                    start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                    queryset = queryset.filter(started_at__gte=start_dt)
                except ValueError:
                    pass
            
            if end_date:
                try:
                    end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                    queryset = queryset.filter(started_at__lte=end_dt)
                except ValueError:
                    pass
            
            # 按时间倒序排列
            queryset = queryset.order_by('-started_at')
            
            # 转换为字典列表
            executions = []
            for exec_obj in queryset:
                executions.append({
                    'id': exec_obj.id,
                    'target': exec_obj.target,
                    'tool_name': exec_obj.tool_name,
                    'analysis_type': exec_obj.analysis_type,
                    'status': exec_obj.status,
                    'started_at': exec_obj.started_at,
                    'finished_at': exec_obj.finished_at,
                    'execution_time': exec_obj.execution_time,
                    'created_by': exec_obj.created_by,
                    'result': exec_obj.result,
                    'error_message': exec_obj.error_message,
                })
            
            if not executions:
                return Response({
                    'error': '没有找到符合条件的执行记录'
                }, status=status.HTTP_404_NOT_FOUND)
            
            exporter = HexStrikeReportExporter()
            
            # 生成文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            target_str = target or 'all'
            filename_base = f'hexstrike_report_{target_str}_{timestamp}'
            
            if format_type == 'excel':
                excel_file = exporter.export_to_excel(executions, target)
                response = HttpResponse(
                    excel_file.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
                return response
            
            elif format_type == 'pdf':
                pdf_file = exporter.export_to_pdf(executions, target)
                response = HttpResponse(pdf_file.read(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
                return response
            
            elif format_type == 'html':
                html_content = exporter.export_to_html(executions, target)
                response = HttpResponse(html_content, content_type='text/html; charset=utf-8')
                response['Content-Disposition'] = f'attachment; filename="{filename_base}.html"'
                return response
            
            else:
                return Response({
                    'error': f'不支持的导出格式: {format_type}。支持格式: excel, pdf, html'
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            logger.exception(f"导出 HexStrike 报告失败: {e}")
            return Response({
                'error': f'导出失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def hexstrike_executions(self, request):
        """
        获取 HexStrike 执行记录列表
        """
        try:
            target = request.query_params.get('target', None)
            status_filter = request.query_params.get('status', None)
            page = int(request.query_params.get('page', 1))
            page_size = int(request.query_params.get('page_size', 20))
            
            queryset = HexStrikeExecution.objects.all()
            
            if target:
                queryset = queryset.filter(target__icontains=target)
            
            if status_filter:
                queryset = queryset.filter(status=status_filter)
            
            queryset = queryset.order_by('-started_at')
            
            # 分页
            total = queryset.count()
            start = (page - 1) * page_size
            end = start + page_size
            executions = queryset[start:end]
            
            # 序列化
            execution_list = []
            for exec_obj in executions:
                execution_list.append({
                    'id': exec_obj.id,
                    'target': exec_obj.target,
                    'tool_name': exec_obj.tool_name,
                    'analysis_type': exec_obj.analysis_type,
                    'status': exec_obj.status,
                    'started_at': exec_obj.started_at.isoformat() if exec_obj.started_at else None,
                    'finished_at': exec_obj.finished_at.isoformat() if exec_obj.finished_at else None,
                    'execution_time': exec_obj.execution_time,
                    'created_by': exec_obj.created_by,
                    'result': exec_obj.result,
                    'error_message': exec_obj.error_message,
                })
            
            return Response({
                'total': total,
                'page': page,
                'page_size': page_size,
                'results': execution_list
            })
            
        except Exception as e:
            logger.exception(f"获取 HexStrike 执行记录失败: {e}")
            return Response({
                'error': f'获取执行记录失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class DingTalkBotView(APIView):
    """钉钉机器人Webhook视图"""
    permission_classes = [AllowAny]  # 钉钉机器人不需要用户认证
    
    def post(self, request):
        """
        接收钉钉机器人的消息（支持Stream推送和Webhook两种方式）
        
        钉钉机器人会发送POST请求到这个接口
        消息格式：
        1. Stream推送格式：
        {
            "headers": {
                "appId": "...",
                "messageId": "...",
                "eventType": "im.message",
                ...
            },
            "body": {
                "content": "用户消息",
                "senderId": "发送者ID",
                ...
            }
        }
        
        2. Webhook格式：
        {
            "msgtype": "text",
            "text": {"content": "用户消息"},
            "senderId": "发送者ID",
            ...
        }
        """
        try:
            # 获取消息数据
            data = request.data
            # 安全记录：不记录可能包含敏感信息的完整数据
            logger.info(f"收到钉钉POST请求: method={request.method}, content_type={request.content_type}, data_size={len(str(data))}")
            
            # 检查是否为Stream推送格式
            is_stream_push = 'headers' in data or 'body' in data
            
            # 从请求头获取应用信息（Stream推送会提供）
            app_id = request.headers.get('X-Dingtalk-App-Id') or request.headers.get('appId')
            agent_code = request.headers.get('X-Dingtalk-Agent-Id') or request.headers.get('agentCode')
            
            # 查找对应的配置
            config = None
            
            if app_id:
                # 优先通过App ID查找配置（Stream推送）
                config = AliyunConfig.objects.filter(
                    dingtalk_enabled=True,
                    dingtalk_app_id=app_id
                ).first()
                logger.info(f"通过App ID查找配置: app_id={app_id}, config={config.id if config else None}")
            
            if not config:
                # 尝试通过Client ID查找（如果有）
                if agent_code:
                    config = AliyunConfig.objects.filter(
                        dingtalk_enabled=True,
                        dingtalk_client_id=agent_code
                    ).first()
                    logger.info(f"通过Agent Code查找配置: agent_code={agent_code}, config={config.id if config else None}")
            
            if not config:
                # 从请求头或URL参数中获取webhook URL（钉钉会提供）
                webhook_url = request.headers.get('X-Webhook-URL') or request.query_params.get('webhook_url')
                
                if webhook_url:
                    # 查找对应的配置（通过webhook URL匹配）
                    config = AliyunConfig.objects.filter(
                        dingtalk_webhook__icontains=webhook_url.split('?')[0],  # 只匹配URL部分，忽略参数
                        dingtalk_enabled=True
                    ).first()
            
            if not config:
                # 如果找不到，使用第一个启用的配置
                if is_stream_push:
                    # Stream推送优先查找有应用凭证的配置
                    config = AliyunConfig.objects.filter(
                        dingtalk_enabled=True,
                        dingtalk_use_stream_push=True
                    ).exclude(
                        dingtalk_client_secret=''
                    ).first()
                
                if not config:
                    config = AliyunConfig.objects.filter(
                        dingtalk_enabled=True
                    ).exclude(
                        dingtalk_webhook=''
                    ).first()
            
            if not config:
                logger.warning("未找到可用的钉钉配置")
                return Response(
                    {
                        'msgtype': 'text',
                        'text': {
                            'content': '❌ 未找到钉钉配置，请在系统配置中添加钉钉应用凭证'
                        }
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            logger.info(f"使用钉钉配置: id={config.id}, name={config.name}, stream_push={config.dingtalk_use_stream_push}")
            
            # 创建钉钉智能体实例
            from app.services.dingtalk_agent import DingTalkAgent
            agent = DingTalkAgent(config.id)
            
            # 获取webhook URL（Stream推送可能不需要）
            webhook_url = config.dingtalk_webhook if config.dingtalk_webhook else None
            
            # 处理消息（异步执行，立即返回）
            # 注意：钉钉要求快速响应（5秒内），所以我们在后台线程中处理
            result = agent.handle_message(webhook_url, data)
            
            # 立即返回确认（钉钉要求快速响应）
            # Stream推送需要返回特定格式
            if is_stream_push:
                return Response({
                    'code': 0,
                    'message': 'success'
                })
            else:
                return Response({
                    'msgtype': 'text',
                    'text': {
                        'content': '消息已接收，正在处理...'
                    }
                })
            
        except Exception as e:
            logger.error(f"处理钉钉机器人消息失败: {e}", exc_info=True)
            # Stream推送需要返回特定错误格式
            is_stream_push = 'headers' in (request.data if hasattr(request, 'data') else {})
            if is_stream_push:
                return Response({
                    'code': 1,
                    'message': f'处理失败: {str(e)}'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            else:
                return Response(
                    {
                        'msgtype': 'text',
                        'text': {
                            'content': f'❌ 处理失败: {str(e)}'
                        }
                    },
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
    
    def get(self, request):
        """
        处理GET请求（用于钉钉Stream推送验证）
        
        根据钉钉官方文档：https://open.dingtalk.com/document/orgapp/configure-stream-push
        钉钉会在配置Stream推送时发送GET请求来验证端点
        
        验证请求会包含以下参数：
        - signature: 签名
        - timestamp: 时间戳
        - nonce: 随机数
        
        需要验证签名，并返回配置信息
        """
        try:
            # 获取验证参数
            signature = request.query_params.get('signature', '')
            timestamp = request.query_params.get('timestamp', '')
            nonce = request.query_params.get('nonce', '')
            echo_str = request.query_params.get('echo_str', '')
            
            logger.info(f"收到钉钉Stream推送验证请求: signature={signature}, timestamp={timestamp}, nonce={nonce}, echo_str={echo_str}")
            
            # 查找启用的钉钉配置（支持Stream推送的配置）
            config = AliyunConfig.objects.filter(
                dingtalk_enabled=True,
                dingtalk_use_stream_push=True
            ).exclude(
                dingtalk_client_secret=''
            ).first()
            
            if not config:
                # 如果没有找到Stream推送配置，尝试查找普通配置
                config = AliyunConfig.objects.filter(
                    dingtalk_enabled=True
                ).exclude(
                    dingtalk_webhook=''
                ).first()
                
                if not config:
                    logger.warning("未找到可用的钉钉配置")
                    return Response({
                        'message': '钉钉机器人接口已就绪，但未找到有效配置',
                        'usage': '请在系统配置中配置钉钉应用凭证（App ID、Client ID、Client Secret）',
                        'endpoint': '/api/dingtalk/bot/'
                    }, status=status.HTTP_200_OK)
            
            # 如果提供了签名验证参数，进行验证
            if signature and timestamp and nonce and config.dingtalk_client_secret:
                # 钉钉Stream推送签名验证逻辑
                # 签名算法：sha256(timestamp + nonce + client_secret)
                sign_str = f"{timestamp}{nonce}{config.dingtalk_client_secret}"
                calculated_sign = hashlib.sha256(sign_str.encode('utf-8')).hexdigest()
                
                if calculated_sign != signature:
                    logger.warning(f"钉钉Stream推送签名验证失败: 计算签名={calculated_sign}, 接收签名={signature}")
                    return Response({
                        'error': '签名验证失败'
                    }, status=status.HTTP_403_FORBIDDEN)
                
                logger.info("钉钉Stream推送签名验证成功")
                
                # 如果提供了echo_str，需要返回它（用于验证）
                if echo_str:
                    return Response({
                        'echo_str': echo_str,
                        'message': '验证成功'
                    })
            
            # 检查是否是checkOnline验证请求（钉钉Stream推送连接检查）
            # 钉钉可能会调用 /checkOnline 端点或带有 checkOnline 参数的请求
            check_online = request.query_params.get('checkOnline') or 'checkOnline' in request.path.lower()
            
            # 获取验证参数（支持多种参数名）
            corp_id = request.query_params.get('corpId') or request.query_params.get('corp_id')
            agent_code = request.query_params.get('agentCode') or request.query_params.get('agent_code') or request.query_params.get('appKey')
            
            logger.info(f"钉钉验证请求详情: path={request.path}, query_params={dict(request.query_params)}, checkOnline={check_online}")
            
            # 处理checkOnline验证请求（钉钉Stream推送连接检查）
            # 这是钉钉用来验证Stream推送连接的主要方式
            # 钉钉Stream推送需要建立WebSocket长连接，checkOnline只是验证端点可用性
            if check_online:
                logger.info(f"收到钉钉Stream推送checkOnline验证请求")
                
                # 检查是否有运行中的Stream服务（通过导入服务模块检查）
                has_running_service = False
                try:
                    from app.services import dingtalk_stream_service
                    if hasattr(dingtalk_stream_service, '_services'):
                        has_running_service = len(dingtalk_stream_service._services) > 0
                        logger.info(f"当前运行中的Stream服务数量: {len(dingtalk_stream_service._services)}")
                except Exception as e:
                    logger.warning(f"检查Stream服务状态时出错: {e}")
                    has_running_service = False
                
                # 查找启用了Stream推送的配置
                stream_configs = AliyunConfig.objects.filter(
                    dingtalk_enabled=True,
                    dingtalk_use_stream_push=True
                ).exclude(
                    dingtalk_client_id=''
                ).exclude(
                    dingtalk_client_secret=''
                )
                
                # 如果提供了agentCode/appKey，尝试匹配
                if agent_code:
                    # 通过Client ID匹配
                    matched_config = stream_configs.filter(dingtalk_client_id=agent_code).first()
                    if matched_config:
                        service_status = False
                        try:
                            from app.services import dingtalk_stream_service
                            if hasattr(dingtalk_stream_service, '_services'):
                                service_status = matched_config.id in dingtalk_stream_service._services
                        except:
                            pass
                        logger.info(f"通过Client ID匹配到配置: {matched_config.id}, 服务运行状态: {service_status}")
                        # 如果有运行中的服务或配置存在，返回成功
                        return Response({
                            'result': True,
                            'success': True
                        })
                    
                    # 通过App ID匹配（备用）
                    matched_config = stream_configs.filter(dingtalk_app_id=agent_code).first()
                    if matched_config:
                        logger.info(f"通过App ID匹配到配置: {matched_config.id}")
                        return Response({
                            'result': True,
                            'success': True
                        })
                
                # 如果有任何启用了Stream推送的配置（已配置Client ID和Secret），返回成功
                # 注意：实际的长连接需要运行 start_dingtalk_stream 命令
                if stream_configs.exists():
                    logger.info(f"找到启用了Stream推送的配置（{stream_configs.count()}个），返回验证成功")
                    return Response({
                        'result': True,
                        'success': True,
                        'message': '配置已就绪，请运行 python manage.py start_dingtalk_stream 启动Stream服务'
                    })
                else:
                    logger.warning("未找到启用了Stream推送的配置")
                    return Response({
                        'result': False,
                        'success': True,
                        'message': '未找到启用了Stream推送的配置，请确保已配置Client ID和Client Secret并启用Stream推送'
                    })
            
            # 处理通过corpId和agentCode的验证请求
            if corp_id and agent_code:
                logger.info(f"收到钉钉配置验证请求: corpId={corp_id}, agentCode={agent_code}")
                
                # 查找匹配的配置
                matched_config = None
                
                # 优先通过Client ID匹配
                if config and config.dingtalk_client_id and config.dingtalk_client_id == agent_code:
                    matched_config = config
                    logger.info(f"通过Client ID匹配到配置: {config.id}")
                # 其次通过App ID匹配
                elif config and config.dingtalk_app_id and config.dingtalk_app_id == agent_code:
                    matched_config = config
                    logger.info(f"通过App ID匹配到配置: {config.id}")
                else:
                    # 尝试在其他配置中查找
                    matched_config = AliyunConfig.objects.filter(
                        Q(dingtalk_client_id=agent_code) | Q(dingtalk_app_id=agent_code),
                        dingtalk_enabled=True
                    ).first()
                    
                    if matched_config:
                        logger.info(f"在其他配置中找到匹配: {matched_config.id}")
                
                if matched_config:
                    return Response({
                        'result': True,
                        'success': True,
                        'code': 0,
                        'message': '验证成功'
                    })
                elif config and config.dingtalk_use_stream_push:
                    # 如果配置了Stream推送但ID不匹配，仍然返回成功（允许后续验证）
                    logger.warning(f"配置存在但ID不匹配，但Stream推送已启用: config_id={config.id}, agentCode={agent_code}")
                    return Response({
                        'result': True,
                        'success': True,
                        'code': 0,
                        'message': '配置存在'
                    })
            
            # 返回配置信息（用于调试和验证）
            response_data = {
                'message': '钉钉机器人接口已就绪',
                'usage': '请使用POST方法发送消息',
                'endpoint': '/api/dingtalk/bot/',
                'stream_push_enabled': config.dingtalk_use_stream_push if config else False,
                'config_id': config.id if config else None,
                'config_name': config.name if config else None,
            }
            
            # 如果配置了应用凭证，显示相关信息（不显示敏感信息）
            if config:
                if config.dingtalk_app_id:
                    response_data['app_id_configured'] = True
                if config.dingtalk_client_id:
                    response_data['client_id_configured'] = True
                if config.dingtalk_client_secret:
                    response_data['client_secret_configured'] = True
                if config.dingtalk_webhook:
                    response_data['webhook_configured'] = True
            
            return Response(response_data)
            
        except Exception as e:
            logger.error(f"处理钉钉Stream推送验证请求失败: {e}", exc_info=True)
            return Response({
                'error': f'处理验证请求失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class FeishuBotView(APIView):
    """
    飞书机器人Webhook视图
    
    注意：
    1. 飞书自定义机器人只能发送消息，不能接收消息
    2. 要实现双向交互，需要使用飞书机器人应用（Bot），并配置事件订阅
    3. 配置步骤：
       a. 在飞书开放平台创建机器人应用
       b. 配置事件订阅，订阅"接收消息"事件
       c. 设置请求地址为: https://your-domain.com/api/feishu/bot/
       d. 在系统配置中添加飞书配置（包含webhook和通义千问AI配置）
    """
    permission_classes = [AllowAny]  # 飞书机器人不需要用户认证
    
    def post(self, request):
        """
        接收飞书机器人的消息
        
        飞书机器人应用会发送POST请求到这个接口
        支持的消息格式：
        1. 事件订阅格式（机器人应用）：
        {
            "schema": "2.0",
            "header": {
                "event_id": "...",
                "event_type": "im.message.receive_v1",
                ...
            },
            "event": {
                "message": {
                    "message_id": "...",
                    "message_type": "text",
                    "content": "{\"text\":\"用户消息\"}"
                },
                "sender": {
                    "sender_id": {
                        "open_id": "...",
                        "user_id": "..."
                    }
                }
            }
        }
        
        2. URL验证（challenge）：
        {
            "challenge": "..."
        }
        """
        try:
            # 获取消息数据
            data = request.data
            
            # 记录收到的请求信息（用于调试）
            logger.info(f"收到飞书机器人请求: method={request.method}, headers={dict(request.headers)}, data={json.dumps(data, ensure_ascii=False)[:1000]}")
            
            # 处理飞书URL验证（如果收到challenge）
            if 'challenge' in data:
                logger.info(f"处理飞书URL验证，challenge={data.get('challenge')}")
                return Response({'challenge': data['challenge']})
            
            # 从请求头或URL参数中获取webhook URL（飞书会提供）
            # 或者从配置中查找对应的webhook
            webhook_url = request.headers.get('X-Webhook-URL') or request.query_params.get('webhook_url')
            
            if not webhook_url:
                # 尝试查找第一个启用的飞书配置
                config = AliyunConfig.objects.filter(
                    feishu_enabled=True
                ).exclude(
                    feishu_webhook=''
                ).first()
                
                if config:
                    webhook_url = config.feishu_webhook
                else:
                    return Response(
                        {
                            'code': 1,
                            'msg': '未找到飞书配置'
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # 查找对应的配置（通过webhook URL匹配）
            config = AliyunConfig.objects.filter(
                feishu_webhook__icontains=webhook_url.split('?')[0],  # 只匹配URL部分，忽略参数
                feishu_enabled=True
            ).first()
            
            if not config:
                # 如果找不到，使用第一个启用的配置
                config = AliyunConfig.objects.filter(
                    feishu_enabled=True
                ).exclude(
                    feishu_webhook=''
                ).first()
            
            if not config:
                return Response(
                    {
                        'code': 1,
                        'msg': '未找到可用的飞书配置'
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # 检查配置是否包含AI配置（智能体需要）
            if not config.qianwen_enabled or not config.qianwen_api_key:
                logger.warning(f"飞书配置 {config.id} 未启用AI或缺少API Key，无法使用智能体功能")
                # 仍然返回成功，但记录警告
                return Response({
                    'code': 0,
                    'msg': 'success',
                    'warning': '配置未启用AI功能，无法使用智能体对话'
                })
            
            # 创建飞书智能体实例
            from app.services.feishu_agent import FeishuAgent
            agent = FeishuAgent(config.id)
            
            # 处理消息（异步执行，立即返回）
            # 注意：飞书要求快速响应（5秒内），所以我们在后台线程中处理
            logger.info(f"开始处理飞书消息，配置ID: {config.id}, webhook: {config.feishu_webhook[:50]}...")
            result = agent.handle_message(config.feishu_webhook, data)
            
            logger.info(f"飞书消息处理结果: {result}")
            
            # 立即返回确认（飞书要求快速响应）
            return Response({
                'code': 0,
                'msg': 'success'
            })
            
        except Exception as e:
            logger.error(f"处理飞书机器人消息失败: {e}", exc_info=True)
            return Response(
                {
                    'code': 1,
                    'msg': f'处理失败: {str(e)}'
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def get(self, request):
        """处理GET请求（用于验证或测试）"""
        # 检查配置
        configs = AliyunConfig.objects.filter(
            feishu_enabled=True
        ).exclude(
            feishu_webhook=''
        )
        
        config_info = []
        for config in configs:
            has_ai = config.qianwen_enabled and bool(config.qianwen_api_key)
            config_info.append({
                'id': config.id,
                'name': config.name,
                'webhook': config.feishu_webhook[:50] + '...' if len(config.feishu_webhook) > 50 else config.feishu_webhook,
                'has_ai_config': has_ai,
                'ai_enabled': config.qianwen_enabled
            })
        
        return Response({
            'message': '飞书机器人接口已就绪',
            'usage': '请使用POST方法发送消息',
            'webhook_url': '配置飞书机器人应用时，将事件订阅地址设置为: https://your-domain.com/api/feishu/bot/',
            'important_note': '⚠️ 飞书自定义机器人只能发送消息，不能接收消息。要实现双向交互，请使用飞书机器人应用（Bot），并配置事件订阅。',
            'configs': config_info,
            'config_count': len(config_info)
        })


class HexStrikeReportDownloadView(APIView):
    """HexStrike 报告下载视图"""
    permission_classes = [AllowAny]  # 报告下载不需要认证，通过链接直接访问

    def get(self, request, filename):
        """
        下载 HexStrike HTML 报告

        Args:
            filename: 报告文件名（如：hexstrike_report_101_37_29_229_20260206_123456.html）
        """
        try:
            from pathlib import Path
            from django.conf import settings

            # 构建 reports 目录路径
            base_dir = Path(settings.BASE_DIR)
            reports_dir = base_dir / 'reports'
            file_path = reports_dir / filename

            # 检查文件是否存在
            if not file_path.exists():
                logger.warning(f"报告文件不存在: {filename}")
                raise Http404(f"报告文件不存在: {filename}")

            # 检查文件名格式（安全检查）
            if not filename.startswith('hexstrike_report_') or not filename.endswith('.html'):
                logger.warning(f"非法的文件名格式: {filename}")
                return Response(
                    {'error': '非法的文件名'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 读取文件内容
            with open(file_path, 'rb') as f:
                file_content = f.read()

            # 返回文件响应
            response = FileResponse(
                io.BytesIO(file_content),
                content_type='text/html; charset=utf-8'
            )
            response['Content-Disposition'] = f'inline; filename="{filename}"'
            response['Content-Length'] = len(file_content)

            logger.info(f"报告下载成功: {filename}, size={len(file_content)} bytes")
            return response

        except Http404:
            raise
        except Exception as e:
            logger.error(f"下载报告失败: {e}", exc_info=True)
            return Response(
                {'error': f'下载失败: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
