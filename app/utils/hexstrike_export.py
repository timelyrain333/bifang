"""
HexStrike 报告导出工具
支持导出 Excel、PDF 和 HTML 格式的报告
"""
import json
import logging
from datetime import datetime
from typing import Dict, Any, List, Optional
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.template.loader import render_to_string
from django.conf import settings
import os

logger = logging.getLogger(__name__)


class HexStrikeReportExporter:
    """HexStrike 报告导出器"""
    
    def __init__(self):
        self.workbook = None
        self.ws = None
    
    def export_to_excel(self, executions: List[Dict[str, Any]], target: Optional[str] = None) -> BytesIO:
        """
        导出执行结果到 Excel
        
        Args:
            executions: HexStrike 执行记录列表
            target: 评估目标（可选，用于文件名）
            
        Returns:
            BytesIO: Excel 文件流
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "工具执行结果"
        
        # 设置表头
        headers = [
            '执行ID', '评估目标', '工具名称', '分析类型', '状态', 
            '开始时间', '结束时间', '执行耗时(秒)', '执行人', '结果摘要'
        ]
        
        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # 写入数据
        row_idx = 2
        for execution in executions:
            result_data = execution.get('result', {})
            result_summary = self._extract_result_summary(result_data)
            
            ws.cell(row=row_idx, column=1, value=execution.get('id', ''))
            ws.cell(row=row_idx, column=2, value=execution.get('target', ''))
            ws.cell(row=row_idx, column=3, value=execution.get('tool_name', '综合分析'))
            ws.cell(row=row_idx, column=4, value=execution.get('analysis_type', 'comprehensive'))
            ws.cell(row=row_idx, column=5, value=self._get_status_display(execution.get('status', '')))
            ws.cell(row=row_idx, column=6, value=self._format_datetime(execution.get('started_at')))
            ws.cell(row=row_idx, column=7, value=self._format_datetime(execution.get('finished_at')))
            ws.cell(row=row_idx, column=8, value=execution.get('execution_time'))
            ws.cell(row=row_idx, column=9, value=execution.get('created_by', ''))
            ws.cell(row=row_idx, column=10, value=result_summary)
            
            # 设置数据行样式
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            row_idx += 1
        
        # 自动调整列宽
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col_idx)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # 创建详细数据工作表
        if executions:
            self._create_detail_sheet(wb, executions)
        
        # 保存到 BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def _create_detail_sheet(self, wb: Workbook, executions: List[Dict[str, Any]]):
        """创建详细数据工作表"""
        ws = wb.create_sheet(title="详细数据")
        
        row_idx = 1
        for execution in executions:
            # 执行信息标题
            title_cell = ws.cell(row=row_idx, column=1, value=f"执行记录 #{execution.get('id', '')} - {execution.get('target', '')}")
            title_cell.font = Font(bold=True, size=12)
            title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            row_idx += 1
            
            # 基本信息
            info = [
                ('评估目标', execution.get('target', '')),
                ('工具名称', execution.get('tool_name', '综合分析')),
                ('分析类型', execution.get('analysis_type', 'comprehensive')),
                ('状态', self._get_status_display(execution.get('status', ''))),
                ('开始时间', self._format_datetime(execution.get('started_at'))),
                ('结束时间', self._format_datetime(execution.get('finished_at'))),
                ('执行耗时', f"{execution.get('execution_time', 0):.2f} 秒" if execution.get('execution_time') else ''),
                ('执行人', execution.get('created_by', '')),
            ]
            
            for label, value in info:
                ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row_idx, column=2, value=str(value))
                row_idx += 1
            
            # 结果数据
            result_data = execution.get('result', {})
            if result_data:
                row_idx += 1
                ws.cell(row=row_idx, column=1, value="执行结果:").font = Font(bold=True)
                row_idx += 1
                
                # 将 JSON 数据格式化为字符串
                result_str = json.dumps(result_data, ensure_ascii=False, indent=2)
                ws.cell(row=row_idx, column=1, value=result_str)
                ws.merge_cells(f'A{row_idx}:B{row_idx}')
                ws.cell(row=row_idx, column=1).alignment = Alignment(vertical="top", wrap_text=True)
                row_idx += 1
            
            # 错误信息
            if execution.get('error_message'):
                row_idx += 1
                ws.cell(row=row_idx, column=1, value="错误信息:").font = Font(bold=True, color="FF0000")
                row_idx += 1
                ws.cell(row=row_idx, column=1, value=execution.get('error_message'))
                ws.merge_cells(f'A{row_idx}:B{row_idx}')
                ws.cell(row=row_idx, column=1).alignment = Alignment(vertical="top", wrap_text=True)
                row_idx += 1
            
            row_idx += 2  # 分隔空行
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
    
    def _extract_result_summary(self, result_data: Dict[str, Any]) -> str:
        """提取结果摘要"""
        if not result_data:
            return ''
        
        # 尝试提取关键信息
        summary_parts = []
        
        if isinstance(result_data, dict):
            # 提取消息
            if 'message' in result_data:
                summary_parts.append(f"消息: {result_data['message']}")
            
            # 提取工具执行结果
            if 'tools' in result_data:
                tools = result_data['tools']
                if isinstance(tools, list):
                    summary_parts.append(f"工具数: {len(tools)}")
            
            # 提取漏洞信息
            if 'vulnerabilities' in result_data:
                vulns = result_data['vulnerabilities']
                if isinstance(vulns, list):
                    summary_parts.append(f"漏洞数: {len(vulns)}")
            
            # 提取端口信息
            if 'ports' in result_data:
                ports = result_data['ports']
                if isinstance(ports, list):
                    summary_parts.append(f"端口数: {len(ports)}")
        
        return ' | '.join(summary_parts) if summary_parts else json.dumps(result_data, ensure_ascii=False)[:100]
    
    def _get_status_display(self, status: str) -> str:
        """获取状态显示文本"""
        status_map = {
            'running': '执行中',
            'success': '成功',
            'failed': '失败',
        }
        return status_map.get(status, status)
    
    def _format_datetime(self, dt) -> str:
        """格式化日期时间"""
        if not dt:
            return ''
        if isinstance(dt, str):
            try:
                dt = datetime.fromisoformat(dt.replace('Z', '+00:00'))
            except:
                return dt
        if isinstance(dt, datetime):
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        return str(dt)
    
    def export_to_html(self, executions: List[Dict[str, Any]], target: Optional[str] = None) -> str:
        """
        导出为 HTML 报告
        
        Args:
            executions: HexStrike 执行记录列表
            target: 评估目标（可选）
            
        Returns:
            str: HTML 内容
        """
        # 计算统计信息
        total_executions = len(executions)
        success_count = sum(1 for e in executions if e.get('status') == 'success')
        failed_count = sum(1 for e in executions if e.get('status') == 'failed')
        running_count = sum(1 for e in executions if e.get('status') == 'running')
        
        # 提取所有工具执行结果
        all_tools = []
        all_vulnerabilities = []
        all_ports = []
        
        for execution in executions:
            result_data = execution.get('result', {})
            if isinstance(result_data, dict):
                # 提取工具结果
                if 'tools' in result_data and isinstance(result_data['tools'], list):
                    for tool in result_data['tools']:
                        if isinstance(tool, dict):
                            tool['execution_id'] = execution.get('id')
                            tool['target'] = execution.get('target')
                            all_tools.append(tool)
                
                # 提取漏洞信息
                if 'vulnerabilities' in result_data and isinstance(result_data['vulnerabilities'], list):
                    for vuln in result_data['vulnerabilities']:
                        if isinstance(vuln, dict):
                            vuln['execution_id'] = execution.get('id')
                            vuln['target'] = execution.get('target')
                            all_vulnerabilities.append(vuln)
                
                # 提取端口信息
                if 'ports' in result_data and isinstance(result_data['ports'], list):
                    for port in result_data['ports']:
                        if isinstance(port, dict):
                            port['execution_id'] = execution.get('id')
                            port['target'] = execution.get('target')
                            all_ports.append(port)
        
        # 格式化执行结果中的 JSON 数据
        formatted_executions = []
        for exec_data in executions:
            formatted_exec = exec_data.copy()
            # 将 result 字段格式化为 JSON 字符串
            if formatted_exec.get('result'):
                formatted_exec['result'] = json.dumps(formatted_exec['result'], ensure_ascii=False, indent=2)
            formatted_executions.append(formatted_exec)
        
        context = {
            'target': target or '综合评估',
            'report_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'total_executions': total_executions,
            'success_count': success_count,
            'failed_count': failed_count,
            'running_count': running_count,
            'executions': formatted_executions,
            'all_tools': all_tools,
            'all_vulnerabilities': all_vulnerabilities,
            'all_ports': all_ports,
        }
        
        # 渲染 HTML 模板
        html_content = render_to_string('hexstrike_report.html', context)
        return html_content
    
    def export_to_pdf(self, executions: List[Dict[str, Any]], target: Optional[str] = None) -> BytesIO:
        """
        导出为 PDF 报告
        
        Args:
            executions: HexStrike 执行记录列表
            target: 评估目标（可选）
            
        Returns:
            BytesIO: PDF 文件流
        """
        try:
            from weasyprint import HTML, CSS
            from django.template.loader import render_to_string
            
            # 生成 HTML
            html_content = self.export_to_html(executions, target)
            
            # 添加 PDF 样式
            pdf_css = CSS(string='''
                @page {
                    size: A4;
                    margin: 2cm;
                }
                body {
                    font-family: "Microsoft YaHei", Arial, sans-serif;
                    font-size: 10pt;
                    line-height: 1.6;
                }
                h1 {
                    color: #366092;
                    border-bottom: 3px solid #366092;
                    padding-bottom: 10px;
                }
                h2 {
                    color: #366092;
                    margin-top: 20px;
                    border-bottom: 1px solid #ccc;
                    padding-bottom: 5px;
                }
                table {
                    width: 100%;
                    border-collapse: collapse;
                    margin: 10px 0;
                }
                th, td {
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                }
                th {
                    background-color: #366092;
                    color: white;
                    font-weight: bold;
                }
                .status-success {
                    color: #67C23A;
                    font-weight: bold;
                }
                .status-failed {
                    color: #F56C6C;
                    font-weight: bold;
                }
                .status-running {
                    color: #E6A23C;
                    font-weight: bold;
                }
            ''')
            
            # 转换为 PDF
            html = HTML(string=html_content)
            pdf_bytes = html.write_pdf(stylesheets=[pdf_css])
            
            output = BytesIO(pdf_bytes)
            output.seek(0)
            return output
            
        except ImportError:
            logger.error("weasyprint 未安装，无法生成 PDF。请运行: pip install weasyprint")
            raise ImportError("PDF 生成功能需要安装 weasyprint 库: pip install weasyprint")